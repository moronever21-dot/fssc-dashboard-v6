"""
Genera ppr_charts_data.js con datos estructurados para Chart.js
Lee directamente los archivos Excel del Drive G:
"""

import json, os, sys, re, time, unicodedata
from pathlib import Path
from collections import defaultdict
import urllib.request
from html.parser import HTMLParser
from datetime import date
sys.path.insert(0, str(Path(__file__).parent.parent))
from sheets_config import PPR_SHEETS
class TableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tables = []
        self._in_td = False
        self._in_th = False
        self._in_tr = False
        self._table = None
        self._row = None
        self._cell = []
    def handle_starttag(self, tag, attrs):
        if tag == 'table':
            self._table = []
        elif tag == 'tr' and self._table is not None:
            self._in_tr = True
            self._row = []
        elif tag in ('td', 'th') and self._in_tr:
            self._in_td = tag == 'td'
            self._in_th = tag == 'th'
            self._cell = []
        elif tag == 'br' and (self._in_td or self._in_th):
            self._cell.append(' ')
    def handle_endtag(self, tag):
        if tag in ('td', 'th') and (self._in_td or self._in_th):
            value = ''.join(self._cell).replace('\xa0', ' ')
            value = re.sub(r'\s+', ' ', value).strip()
            self._row.append(value)
            self._in_td = False
            self._in_th = False
            self._cell = []
        elif tag == 'tr' and self._in_tr:
            if self._row and any(cell for cell in self._row):
                self._table.append(self._row)
            self._row = None
            self._in_tr = False
        elif tag == 'table' and self._table is not None:
            self.tables.append(self._table)
            self._table = None
    def handle_data(self, data):
        if self._in_td or self._in_th:
            self._cell.append(data)

def fetch_text(url):
    last_error = None
    for attempt in range(4):
        try:
            with urllib.request.urlopen(url, timeout=45) as response:
                return response.read().decode('utf-8', 'ignore')
        except Exception as error:
            last_error = error
            time.sleep(1.2 * (attempt + 1))
    raise last_error

def get_tabs(sheet_id):
    text = fetch_text(f'https://docs.google.com/spreadsheets/d/{sheet_id}/htmlview')
    pattern = re.compile(r'items.push\(\{name: "([^"]+)", pageUrl: "([^"]+)", gid: "([^"]+)"')
    tabs = []
    for name, page_url, gid in pattern.findall(text):
        tabs.append({
            'name': name,
            'gid': gid,
            'pageUrl': page_url.replace('\\/', '/').replace('\\x3d', '=')
        })
    return tabs

def get_table(sheet_id, gid):
    text = fetch_text(f'https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:html&gid={gid}')
    parser = TableParser()
    parser.feed(text)
    return parser.tables[0] if parser.tables else []

def to_float(value):
    text = str(value).strip().replace('%', '').replace(',', '.')
    text = text.replace('<', '').replace('>', '').replace('~', '')
    if not text:
        return None
    text = re.sub(r'[^0-9.\-]', '', text)
    if text.count('.') > 1:
        parts = text.split('.')
        text = ''.join(parts[:-1]) + '.' + parts[-1]
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None

def mean(values):
    valid = [value for value in values if isinstance(value, (int, float))]
    if not valid:
        return None
    return round(sum(valid) / len(valid), 2)


# Forzar salida UTF-8
sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl no instalado")
    sys.exit(1)

PPR_ROOT = Path("G:\\Mi unidad\\FSSC_22000_Automatizacion\\Programas de Prerrequisitos (PPR)")

output = {
    "residuos": {},
    "bpm_lavado": {},
    "quebradizos": {},
    "resumen": {},
    "limpieza_desinfeccion": {},
    "higiene_personal": {},
    "control_plagas": {}
}

# =====================================================
# 1. ELIMINACIÓN DE RESIDUOS
# =====================================================
print("=== Procesando Eliminación de Residuos ===")
residuos_path = PPR_ROOT / "Eliminación de Residuos y Desperdicios"

residuos_por_area = defaultdict(lambda: {"chias": [], "podridas": [], "fechas": []})
totales_por_fecha = defaultdict(lambda: {"chias": 0.0, "podridas": 0.0})

for fpath in residuos_path.glob("*.xlsm"):
    print(f"  Leyendo: {fpath.name}")
    try:
        wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
        if "Datos " in wb.sheetnames or "Datos" in wb.sheetnames:
            sheet_name = "Datos " if "Datos " in wb.sheetnames else "Datos"
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            header_idx = None
            for i, row in enumerate(rows):
                if row and str(row[0]).strip() in ["Nº IA", "N° IA", "IA"]:
                    header_idx = i
                    break
            if header_idx is None:
                # Try to find data rows by pattern
                for i, row in enumerate(rows):
                    if row and row[0] and str(row[0]).startswith("IA-"):
                        header_idx = i - 1
                        break
            
            for row in rows[1:]:  # skip header
                if not row or not row[0]:
                    continue
                ia = str(row[0]) if row[0] else ""
                if not ia.startswith("IA"):
                    continue
                fecha = row[1]
                area = str(row[2]).strip() if row[2] else ""
                
                # kg chias
                chias = None
                try:
                    chias = float(row[3]) if row[3] else None
                except:
                    pass
                
                # kg podridas
                podridas = None
                try:
                    podridas = float(row[4]) if len(row) > 4 and row[4] else None
                except:
                    pass
                
                if fecha and area:
                    fecha_str = str(fecha)[:10]
                    if chias:
                        residuos_por_area[area]["fechas"].append(fecha_str)
                        residuos_por_area[area]["chias"].append(chias)
                        totales_por_fecha[fecha_str]["chias"] += chias
                    if podridas:
                        residuos_por_area[area]["podridas"].append(podridas)
                        totales_por_fecha[fecha_str]["podridas"] += podridas
        wb.close()
    except Exception as e:
        print(f"    ERROR: {e}")

# Ordenar por fecha
print(f"  Areas encontradas: {list(residuos_por_area.keys())}")
print(f"  Fechas únicas: {sorted(totales_por_fecha.keys())}")

output["residuos"] = {
    "por_area": {k: v for k, v in residuos_por_area.items()},
    "totales_por_fecha": {k: v for k, v in sorted(totales_por_fecha.items())},
    "areas": list(residuos_por_area.keys())
}

# =====================================================
# 2. BPM - RC.HP.05 LAVADO DE ROPA
# =====================================================
print("\n=== Procesando BPM - Lavado de Ropa ===")
bpm_path = PPR_ROOT / "BPM 2026"

lavado_rows = []
for fpath in bpm_path.glob("RC.HP.05*.xlsx"):
    print(f"  Leyendo: {fpath.name}")
    try:
        wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
        for sname in wb.sheetnames:
            if "Lavado" in sname or "HP.05" in sname:
                ws = wb[sname]
                rows = list(ws.iter_rows(values_only=True))
                
                # Encontrar fila de datos (tiene fechas)
                data_start = None
                for i, row in enumerate(rows):
                    if row and row[0] and hasattr(row[0], 'year'):
                        data_start = i
                        break
                    if row and row[0] and '2026' in str(row[0]):
                        data_start = i
                        break
                
                if data_start is None:
                    print(f"    No se encontró inicio de datos en {sname}")
                    continue
                
                for row in rows[data_start:]:
                    if not row or not row[0]:
                        continue
                    fecha = row[0]
                    if not fecha or (not hasattr(fecha,'year') and '2026' not in str(fecha) and '2025' not in str(fecha)):
                        continue
                    
                    try:
                        # Cols: fecha_entrega, hora, responsable_entrega, area, codigo, casaca, pantalon, cofia, barbijo, responsable_lavado, fecha_devolucion, hora_dev, cantidad_dev, codigo_dev, responsable_dev
                        fecha_str = str(fecha)[:10]
                        area = str(row[3]).strip() if row[3] else ""
                        casaca = float(row[5]) if row[5] and str(row[5]) not in ['_',''] else 0
                        pantalon = float(row[6]) if row[6] and str(row[6]) not in ['_',''] else 0
                        cofia = float(row[7]) if row[7] and str(row[7]) not in ['_',''] else 0
                        barbijo = float(row[8]) if row[8] and str(row[8]) not in ['_',''] else 0
                        responsable = str(row[2]).strip() if row[2] else ""
                        cantidad_dev = None
                        try:
                            cantidad_dev = float(row[12]) if len(row) > 12 and row[12] and str(row[12]) not in ['_',''] else None
                        except:
                            pass
                        
                        total_prendas = casaca + pantalon + cofia + barbijo
                        if total_prendas > 0 or (casaca + pantalon > 0):
                            lavado_rows.append({
                                "fecha": fecha_str,
                                "area": area,
                                "responsable": responsable,
                                "casaca": int(casaca),
                                "pantalon": int(pantalon),
                                "cofia": int(cofia),
                                "barbijo": int(barbijo),
                                "total": int(total_prendas),
                                "cant_devuelta": int(cantidad_dev) if cantidad_dev else None
                            })
                    except Exception as e:
                        pass  # Skip bad rows
        wb.close()
    except Exception as e:
        print(f"    ERROR: {e}")

print(f"  Registros de lavado: {len(lavado_rows)}")
for r in lavado_rows[:10]:
    print(f"    {r}")

# Agrupar por área
lavado_por_area = defaultdict(lambda: {"total":0,"casaca":0,"pantalon":0,"cofia":0,"barbijo":0,"registros":0})
lavado_por_fecha = defaultdict(lambda: {"total":0,"registros":0})

for r in lavado_rows:
    a = r["area"] if r["area"] else "Sin área"
    lavado_por_area[a]["total"] += r["total"]
    lavado_por_area[a]["casaca"] += r["casaca"]
    lavado_por_area[a]["pantalon"] += r["pantalon"]
    lavado_por_area[a]["cofia"] += r["cofia"]
    lavado_por_area[a]["barbijo"] += r["barbijo"]
    lavado_por_area[a]["registros"] += 1
    
    f = r["fecha"]
    lavado_por_fecha[f]["total"] += r["total"]
    lavado_por_fecha[f]["registros"] += 1

output["bpm_lavado"] = {
    "registros": lavado_rows,
    "por_area": {k: v for k, v in lavado_por_area.items()},
    "por_fecha": {k: v for k, v in sorted(lavado_por_fecha.items())},
    "total_prendas": sum(r["total"] for r in lavado_rows),
    "total_registros": len(lavado_rows)
}

# =====================================================
# 3. MATERIALES QUEBRADIZOS - Inventario y Riesgo
# =====================================================
print("\n=== Procesando Materiales Quebradizos ===")
mq_root = PPR_ROOT / "materiales quebradizos"

riesgo_data = []  # {area, material, tipo_mat, impacto, probabilidad, riesgo_total, frecuencia}
inventario_data = []  # {area, tipo_mat, elemento, cantidad, estado_aceptable, estado_inaceptable}
incidentes_data = []

def read_quebradizo(fpath, tipo_mat):
    """Lee un archivo de quebradizos y extrae riesgo e inventario.
    Estructura real (todas las hojas): col A (idx 0) siempre None, datos desde col B (idx 1).
    Riesgo (RC.GV.02/GM.02/MM.02): B=No, C/D/E=Elemento, F=Impacto, G=Prob, H=RiesgoTotal, L=Frecuencia
    Inventario (RC.GV.01/GM.01/MM.01): B/C/D=Elemento, G=Cantidad, H=Aceptable, I=Inaceptable
    """
    import re
    local_riesgo = []
    local_inv = []
    try:
        wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)

        # Nombre de área por defecto desde el nombre del archivo
        area_nombre = fpath.stem
        m = re.search(r'\d+[_\s](.+)$', fpath.stem)
        if m:
            area_nombre = m.group(1).strip()

        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))

            # Extraer nombre del área de la fila que contiene "ÁREA:" (fila 8, idx 7)
            # row[1] = 'ÁREA: Recepción y Almacenamiento de MPH'
            for row in rows[:15]:
                if not row:
                    continue
                for c in row:
                    if c and 'ÁREA' in str(c) and ':' in str(c):
                        area_nombre = str(c).split(':', 1)[1].strip().split('\n')[0].strip()
                        break

            if 'Riesgo' in sname or 'GV.02' in sname or 'GM.02' in sname or 'MM.02' in sname:
                # Hoja de riesgo: filas de datos tienen row[1] = número (1,2,3...)
                for row in rows:
                    if not row or row[1] is None:
                        continue
                    try:
                        num = int(float(row[1]))
                        if num < 1 or num > 200:
                            continue
                    except (TypeError, ValueError):
                        continue
                    # Elemento: primero de col C/D/E (idx 2/3/4) que no sea '-'
                    elemento = ""
                    for c in (row[2], row[3], row[4]) if len(row) > 4 else row[2:]:
                        if c and str(c).strip() not in ['-', '', 'None']:
                            elemento = str(c).strip()[:40]
                            break
                    if not elemento:
                        continue
                    try:
                        impacto = float(row[5]) if len(row) > 5 and row[5] is not None else 0
                        probabilidad = float(row[6]) if len(row) > 6 and row[6] is not None else 0
                        riesgo = float(row[7]) if len(row) > 7 and row[7] is not None else round(impacto * probabilidad, 1)
                    except (TypeError, ValueError):
                        continue
                    if impacto == 0 and probabilidad == 0:
                        continue
                    # Frecuencia en col L (idx 11)
                    frecuencia = "MENSUAL"
                    if len(row) > 11 and row[11] and str(row[11]).upper() in ['SEMANAL', 'QUINCENAL', 'MENSUAL']:
                        frecuencia = str(row[11]).upper()
                    local_riesgo.append({
                        "area": area_nombre[:50],
                        "tipo_mat": tipo_mat,
                        "elemento": elemento,
                        "impacto": impacto,
                        "probabilidad": probabilidad,
                        "riesgo": riesgo,
                        "frecuencia": frecuencia
                    })

            elif 'Relevamiento' in sname or 'GV.01' in sname or 'GM.01' in sname or 'MM.01' in sname:
                # Hoja de inventario: elemento en col B/C/D (idx 1/2/3), cant en col G (idx 6)
                SKIP_WORDS = {'-', '', 'none', 'vidrio', 'plástico', 'plastico',
                              'cerámica', 'ceramica', 'madera', 'metal', 'elemento',
                              'cantidad', 'estado', 'correcciones', 'observaciones'}
                for row in rows:
                    if not row:
                        continue
                    # Elemento: primero de idx 1/2/3 que sea texto real
                    elemento = None
                    for c in (row[1], row[2], row[3]) if len(row) > 3 else row[1:]:
                        s = str(c).strip() if c is not None else ''
                        if s and s.lower() not in SKIP_WORDS and len(s) > 2:
                            elemento = s[:40]
                            break
                    if not elemento:
                        continue
                    try:
                        cantidad = float(row[6]) if len(row) > 6 and row[6] is not None else None
                        if cantidad is None:
                            continue
                        aceptable = float(row[7]) if len(row) > 7 and row[7] is not None else 0
                        inaceptable = float(row[8]) if len(row) > 8 and row[8] is not None else 0
                    except (TypeError, ValueError):
                        continue
                    local_inv.append({
                        "area": area_nombre[:50],
                        "tipo_mat": tipo_mat,
                        "elemento": elemento,
                        "cantidad": int(cantidad),
                        "aceptable": int(aceptable),
                        "inaceptable": int(inaceptable)
                    })
        wb.close()
    except Exception as e:
        print(f"    ERROR en {fpath.name}: {e}")
    return local_riesgo, local_inv

for subcarpeta, tipo_mat in [("PLASTICO VIDRIO CERAMICA","VPC"), ("MADERA","MAD"), ("METALES","MET")]:  # nombres reales del Drive
    sub_path = mq_root / subcarpeta
    if not sub_path.exists():
        print(f"  No existe: {sub_path}")
        continue
    archivos = list(sub_path.glob("*.xlsx"))
    print(f"  {subcarpeta}: {len(archivos)} archivos")
    for fpath in archivos:
        r, inv = read_quebradizo(fpath, tipo_mat)
        riesgo_data.extend(r)
        inventario_data.extend(inv)
        print(f"    {fpath.name}: {len(r)} riesgos, {len(inv)} inventario")

print(f"\n  Total riesgos: {len(riesgo_data)}")
print(f"  Total inventario: {len(inventario_data)}")

# Agrupar riesgo por área (máximo por área)
riesgo_por_area = defaultdict(lambda: {"VPC":0,"MAD":0,"MET":0,"max":0,"elementos":0})
for r in riesgo_data:
    a = r["area"]
    t = r["tipo_mat"]
    rv = r["riesgo"]
    if rv > riesgo_por_area[a][t]:
        riesgo_por_area[a][t] = rv
    if rv > riesgo_por_area[a]["max"]:
        riesgo_por_area[a]["max"] = rv
    riesgo_por_area[a]["elementos"] += 1

output["quebradizos"] = {
    "riesgo_detalle": riesgo_data[:100],
    "inventario": inventario_data[:100],
    "riesgo_por_area": {k: v for k, v in riesgo_por_area.items()},
    "total_elementos_inventario": len(inventario_data),
    "total_riesgos": len(riesgo_data)
}


# =====================================================
# 4. CONTROL PLAGAS, LIMPIEZA, HIGIENE PERSONAL (Google Sheets)
# =====================================================
print("\n=== Procesando módulos Google Sheets ===")
plagas_id = PPR_SHEETS.get("CONTROL_PLAGAS")
limpieza_id = PPR_SHEETS.get("LIMPIEZA_DESINFECCION")
higiene_id = PPR_SHEETS.get("HIGIENE_PERSONAL")

# --- CONTROL PLAGAS ---
plagas_data = {}
plagas_incidentes = 0
if plagas_id:
    try:
        tabs = get_tabs(plagas_id)
        # Buscar tab con "Incidentes" o similar
        tab = next((t for t in tabs if "incidente" in t['name'].lower() or "plaga" in t['name'].lower()), tabs[0])
        rows = get_table(plagas_id, tab['gid'])
        # Buscar columnas relevantes
        headers = [c.strip().lower() for c in rows[0]] if rows else []
        zona_idx = next((i for i, h in enumerate(headers) if "zona" in h), None)
        estado_idx = next((i for i, h in enumerate(headers) if "estado" in h or "situacion" in h), None)
        fecha_idx = next((i for i, h in enumerate(headers) if "fecha" in h), None)
        plagas_incidentes = 0
        plagas_rows = []
        for row in rows[1:]:
            if not row or (estado_idx is not None and row[estado_idx] and str(row[estado_idx]).strip().lower() in ["cerrado", "solucionado", "resuelto"]):
                continue
            plagas_incidentes += 1
            plagas_rows.append({
                "zona": row[zona_idx] if zona_idx is not None else "",
                "estado": row[estado_idx] if estado_idx is not None else "",
                "fecha": row[fecha_idx] if fecha_idx is not None else ""
            })
        plagas_data = {
            "incidentes": plagas_incidentes,
            "detalles": plagas_rows[:30],
            "total_registros": len(rows)-1
        }
    except Exception as e:
        plagas_data = {"error": str(e)}

# --- LIMPIEZA Y DESINFECCIÓN ---
limpieza_data = {}
limpieza_cumplimiento = None
if limpieza_id:
    try:
        tabs = get_tabs(limpieza_id)
        # Buscar tab con "cumplimiento" o "cronograma" o usar la primera
        tab = next((t for t in tabs if "cumplimiento" in t['name'].lower() or "cronograma" in t['name'].lower()), tabs[0])
        rows = get_table(limpieza_id, tab['gid'])
        # Buscar columna con "%" o "cumplimiento"
        headers = [c.strip().lower() for c in rows[0]] if rows else []
        pct_idx = next((i for i, h in enumerate(headers) if "%" in h or "cumplimiento" in h), None)
        valores = [to_float(row[pct_idx]) for row in rows[1:] if pct_idx is not None and row[pct_idx]]
        limpieza_cumplimiento = mean([v for v in valores if v is not None])
        limpieza_data = {
            "cumplimiento": limpieza_cumplimiento,
            "historial": valores[-30:],
            "total_registros": len(valores)
        }
    except Exception as e:
        limpieza_data = {"error": str(e)}

# --- HIGIENE PERSONAL ---
higiene_data = {}
higiene_cumplimiento = None
if higiene_id:
    try:
        tabs = get_tabs(higiene_id)
        # Buscar tab con "cumplimiento" o "chequeo" o usar la primera
        tab = next((t for t in tabs if "cumplimiento" in t['name'].lower() or "chequeo" in t['name'].lower()), tabs[0])
        rows = get_table(higiene_id, tab['gid'])
        headers = [c.strip().lower() for c in rows[0]] if rows else []
        pct_idx = next((i for i, h in enumerate(headers) if "%" in h or "cumplimiento" in h), None)
        valores = [to_float(row[pct_idx]) for row in rows[1:] if pct_idx is not None and row[pct_idx]]
        higiene_cumplimiento = mean([v for v in valores if v is not None])
        higiene_data = {
            "cumplimiento": higiene_cumplimiento,
            "historial": valores[-30:],
            "total_registros": len(valores)
        }
    except Exception as e:
        higiene_data = {"error": str(e)}

output["control_plagas"] = plagas_data
output["limpieza_desinfeccion"] = limpieza_data
output["higiene_personal"] = higiene_data

# =====================================================
# RESUMEN GENERAL Y KPIs
# =====================================================
total_residuos_kg = sum(totales_por_fecha.get(f,{}).get("chias",0) for f in totales_por_fecha)
areas_residuos = len(residuos_por_area)
registros_lavado = len(lavado_rows)
total_prendas = sum(r["total"] for r in lavado_rows)
total_elementos_qb = len(inventario_data)
areas_alto_riesgo = sum(1 for a in riesgo_por_area.values() if a["max"] >= 6)


# Forzar claves requeridas por el dashboard aunque estén vacías
for key in ["bpm_lavado", "limpieza_desinfeccion", "higiene_personal", "control_plagas"]:
    if key not in output or not isinstance(output[key], dict):
        output[key] = {}

# KPIs nuevos
modulos_cumplimiento = [v for v in [limpieza_cumplimiento, higiene_cumplimiento] if v is not None]
if registros_lavado > 0:
    modulos_cumplimiento.append(100)  # Si hay registros de lavado, se asume cumplimiento 100%
if total_residuos_kg > 0:
    modulos_cumplimiento.append(100)
if total_elementos_qb > 0:
    modulos_cumplimiento.append(100)
if plagas_incidentes == 0:
    modulos_cumplimiento.append(100)

cumplimiento_general = mean(modulos_cumplimiento) if modulos_cumplimiento else None
indice_higiene = mean([v for v in [higiene_cumplimiento] if v is not None])
eficacia_limpieza = limpieza_cumplimiento
estado_plagas = plagas_incidentes

output["resumen"] = {
    "total_residuos_kg": round(total_residuos_kg, 1),
    "areas_residuos": areas_residuos,
    "registros_lavado": registros_lavado,
    "total_prendas_lavadas": total_prendas,
    "total_elementos_quebradizos": total_elementos_qb,
    "areas_alto_riesgo": areas_alto_riesgo,
    "modulos_activos": 5,
    "cumplimiento_general": cumplimiento_general,
    "indice_higiene": indice_higiene,
    "eficacia_limpieza": eficacia_limpieza,
    "estado_plagas": estado_plagas
}

# =====================================================
# GUARDAR SALIDA
# =====================================================
# Ruta relativa: PPR/ -> raiz -> PPRo/ppr_charts_data.js
out_path = Path(__file__).parent.parent / "PPRo" / "ppr_charts_data.js"
with open(out_path, "w", encoding="utf-8") as f:
    f.write("// Datos generados automáticamente para gráficos PPR FSSC 22000\n")
    f.write("// Generado: " + __import__('datetime').datetime.now().strftime("%Y-%m-%d %H:%M") + "\n")
    f.write("const PPR_CHARTS_DATA = ")
    json.dump(output, f, ensure_ascii=False, indent=2)
    f.write(";\n")

print(f"\n=== ARCHIVO GENERADO: {out_path} ===")
print(f"Resumen: {output['resumen']}")
