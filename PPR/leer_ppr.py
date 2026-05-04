import openpyxl, os, json

base = r'G:\Mi unidad\FSSC_22000_Automatizacion\Programas de Prerrequisitos (PPR)'

result = {}

def leer_xlsx(filepath, nombre, max_rows=15):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        hojas = []
        for sh in wb.sheetnames:
            ws = wb[sh]
            filas = []
            for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
                vals = [str(v) for v in row if v is not None]
                if vals:
                    filas.append(vals)
            hojas.append({"hoja": sh, "filas": filas, "total_filas": ws.max_row})
        return {"archivo": nombre, "hojas": hojas}
    except Exception as e:
        return {"archivo": nombre, "error": str(e)}

# === BPM 2026 ===
bpm_path = os.path.join(base, "BPM 2026")
bpm_archivos = []
for f in os.listdir(bpm_path):
    if f.endswith(".xlsx"):
        bpm_archivos.append(leer_xlsx(os.path.join(bpm_path, f), f))
result["BPM_2026"] = bpm_archivos

# === Eliminacion de Residuos ===
elim_path = os.path.join(base, "Eliminación de Residuos y Desperdicios")
elim_archivos = []
for f in os.listdir(elim_path):
    if f.lower().endswith(".xlsx") or f.lower().endswith(".xlsm") or f.lower().endswith(".xls"):
        elim_archivos.append(leer_xlsx(os.path.join(elim_path, f), f))
result["Eliminacion_Residuos"] = elim_archivos

# === Materiales Quebradizos - leer 1 de cada subcarpeta ===
mat_base = os.path.join(base, "materiales quebradizos")
mat_result = {}
if os.path.exists(mat_base):
    for subfolder in os.listdir(mat_base):
        sub_path = os.path.join(mat_base, subfolder)
        if os.path.isdir(sub_path):
            archivos_sub = []
            for f in sorted(os.listdir(sub_path))[:2]:  # solo primeros 2
                if f.lower().endswith(".xlsx"):
                    archivos_sub.append(leer_xlsx(os.path.join(sub_path, f), f, max_rows=12))
            mat_result[subfolder] = archivos_sub
result["Materiales_Quebradizos"] = mat_result

# === Material de Embalaje ===
emb_path = os.path.join(base, "Material de Embalaje")
# los .gsheet no se pueden leer directamente - son accesos directos
result["Material_Embalaje"] = [f for f in os.listdir(emb_path)]

# === Control Plagas ===
plg_path = os.path.join(base, "Control Plagas")
result["Control_Plagas"] = [f for f in os.listdir(plg_path)]

# Guardar resultado
with open("ppr_data.json", "w", encoding="utf-8") as fout:
    json.dump(result, fout, ensure_ascii=False, indent=2)

print("DONE - ppr_data.json generado")
print("BPM archivos:", len(result["BPM_2026"]))
print("Eliminacion archivos:", len(result["Eliminacion_Residuos"]))
print("Control Plagas:", result["Control_Plagas"])
print("Material Embalaje:", result["Material_Embalaje"])
